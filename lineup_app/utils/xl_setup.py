from openpyxl import load_workbook
import os
import numpy as np



def read_conns(well_data):
    # get current directory using os library
    # dirname, filename = os.path.split(os.path.abspath(__file__))
    dirname=os.path.abspath(os.path.join(os.path.dirname( __file__ ), '..')) # added after restructuring files/folders
    # construct excel file full path
    xl_fullpath=os.path.join(dirname,'static\well_connections.xlsm')
    # load excel file
    conn_wb = load_workbook(xl_fullpath)
    # load sheet
    conns=conn_wb.get_sheet_by_name(name='connections2')
    # get max row number. Needed for furture looping through sheet rows
    row_cnt = conns.max_row

    # init vars
    well_details={}
    well=""
    well_conns={}
    # start from row 3 in excel
    r=3
    while conns.cell(row=r, column=1).value!=None: # loop through rows until empty cell is met

        # if well is not wellname in excel, then new well data has started
        # then remember prev well, init all vars for the new well
        if well!=str(conns.cell(row=r, column=1).value):

            if well: # to save prev well data
                well_details[well]=well_conns

            # init all vars for the new well
            well_conns={}
            well=str(conns.cell(row=r, column=1).value)
            well_conns["wellname"]=well
            well_conns["connected"]=conns.cell(row=r, column=2).value
            well_conns["routes"]=[]

        # remember TL if exists, make none if not
        if conns.cell(row=r, column=6).value!=None:
            tl=str(conns.cell(row=r, column=6).value)
        else:
            tl=""

        # check if comingled columns have wells
        if conns.cell(row=r, column=9).value==None: # first check colunm 9
            comingled=""
        else:
            if conns.cell(row=r, column=10).value==None: # then check colunm 10
                comingled=str(conns.cell(row=r, column=9).value)
            else: # both 9 and 10 have wells
                # concatenate 2 comingled through comma
                comingled=str(conns.cell(row=r, column=9).value)+", "+str(conns.cell(row=r, column=10).value)

        if conns.cell(row=r, column=7).value!=None:
            os_str=str(conns.cell(row=r, column=7).value)
        else:
            os_str=""

        # 11th colunm temporarily
        if conns.cell(row=r, column=11).value!=None:
            fl_pipe_os=str(conns.cell(row=r, column=11).value)
        else:
            fl_pipe_os=""


        well_conns["routes"].append({
            "unit":str(conns.cell(row=r, column=3).value),
            "rms":str(conns.cell(row=r, column=4).value),
            "slot":str(conns.cell(row=r, column=5).value),
            "tl":tl,
            "os":os_str,
            "default":str(conns.cell(row=r, column=8).value),
            # "comingled":comingled,
            "fl_pipe_os":fl_pipe_os,
            "route_name":str(conns.cell(row=r, column=3).value)+"--"+str(conns.cell(row=r, column=4).value)+"--"+tl+"--slot "+str(conns.cell(row=r, column=5).value)
        })
        well_conns["comingled"]=comingled

        r+=1


    if well: # to save very last well
        well_details[well]=well_conns



    for well,m in well_data.items(): # merge with well_data
        if well in well_details:
            # print(well,well_details[well])
            well_data[well]["connection"]=well_details[well] # assign list of connections

    return well_data


def read_pcs():
    # dirname, filename = os.path.split(os.path.abspath(__file__))
    dirname=os.path.abspath(os.path.join(os.path.dirname( __file__ ), '..')) # added after restructuring files/folders
    xl_fullpath=os.path.join(dirname,'static\Deliverability.xlsx')
    pc_wb = load_workbook(xl_fullpath)
    pc_sh=pc_wb.get_sheet_by_name(name='Updated data')

    well_pcs={}

    r=8
    while pc_sh.cell(row=r, column=2).value!=None:
        if pc_sh.cell(row=r, column=16).value!=None:
            well_data={}
            well_data["wellname"]=str(pc_sh.cell(row=r, column=2).value)
            well_data["sbhp"]=pc_sh.cell(row=r, column=12).value
            if pc_sh.cell(row=r, column=14).value!=None and pc_sh.cell(row=r, column=14).value>0:
                well_data["gor"]=(1.0/pc_sh.cell(row=r, column=14).value)*1000000.0
            else:
                well_data["gor"]=0.0

            if pc_sh.cell(row=r, column=32).value!=None:
                well_data["map"]=pc_sh.cell(row=r, column=32).value
            else:
                well_data["map"]=0.0

            if pc_sh.cell(row=r, column=8).value!=None:
                well_data["wct"]=pc_sh.cell(row=r, column=8).value/100.0
            else:
                well_data["wct"]=0.0

            if well_data["gor"]>0.0 and well_data["wct"]>0.0:
                well_data["wgr"]=well_data["wct"]/(well_data["gor"]*(1.0-well_data["wct"]))
            else:
                well_data["wgr"]=0.0


            well_data["qoil_coeffs"]=[
                pc_sh.cell(row=r, column=16).value,\
                pc_sh.cell(row=r, column=17).value,\
                pc_sh.cell(row=r, column=18).value,\
                pc_sh.cell(row=r, column=19).value
            ]

            well_data["fbhp_coeffs"]=[
                pc_sh.cell(row=r, column=20).value,\
                pc_sh.cell(row=r, column=21).value,\
                pc_sh.cell(row=r, column=22).value,\
                pc_sh.cell(row=r, column=23).value
            ]

            well_pcs[well_data["wellname"]]=well_data
        r+=1

    for well,data in well_pcs.items():
        pc=generate_pc(data["qoil_coeffs"],data["wct"],data["fbhp_coeffs"],data["gor"])
        well_pcs[well]["pc"]=pc


    return well_pcs

def generate_pc(qoil_coeffs,wc,fbhp_coeffs,gor):
    qoil=1000.0
    thp=50.0
    while qoil>0.05:
        qoil=qoil_coeffs[0]*thp**3.0+qoil_coeffs[1]*thp**2.0+qoil_coeffs[2]*thp+qoil_coeffs[3]
        thp+=0.001
    thps=np.linspace(50.0,thp,20.0)
    qoil=qoil_coeffs[0]*thps**3.0+qoil_coeffs[1]*thps**2.0+qoil_coeffs[2]*thps+qoil_coeffs[3]
    qliqs=qoil/(1.0-wc)
    qgas=gor*qoil/1000.0
    fbhps=fbhp_coeffs[0]*thps**3.0+fbhp_coeffs[1]*thps**2.0+fbhp_coeffs[2]*thps+fbhp_coeffs[3]
    temps=qliqs*0.01842105263+13.15789473684

    # null last points to avoid small values when well is shut-in.
    qliqs[-1]=0.0
    qgas[-1]=0.0

    return {"thps":thps.tolist(),"qliqs":qliqs.tolist(),"fbhps":fbhps.tolist(),"qgas":qgas.tolist(),"temps":temps.tolist()}



def read_maps(well_data):
    # dirname, filename = os.path.split(os.path.abspath(__file__))
    dirname=os.path.abspath(os.path.join(os.path.dirname( __file__ ), '..')) # added after restructuring files/folders
    xl_fullpath=os.path.join(dirname,'static\Deliverability.xlsx')
    pc_wb = load_workbook(xl_fullpath)
    pc_sh=pc_wb.get_sheet_by_name(name='Updated data')

    well_maps={}

    r=8
    while pc_sh.cell(row=r, column=2).value!=None:
        if pc_sh.cell(row=r, column=16).value!=None:
            data={}
            data["wellname"]=str(pc_sh.cell(row=r, column=2).value)

            if pc_sh.cell(row=r, column=32).value!=None:
                data["map"]=pc_sh.cell(row=r, column=32).value
            else:
                data["map"]=0.0
            well_maps[data["wellname"]]=data
        r+=1

    for well,m in well_data.items(): # merge with well_data
        if well in well_maps:
            # print(well,well_maps[well]["map"])
            well_data[well]["map"]=well_maps[well]["map"]

    return well_data



def make_well_data_by_unit(session_json):
    # group wells by unit for html page
    session_json["well_data_byunit"]=[{},{},{}]
    for well,val in session_json["well_data"].items():
        if "masked" in val: #required to know which well is masked/unmasked in GAP to include/exclude well in the list
            if val["masked"]==0:
                session_json["well_data_byunit"][val["unit_id"]][well]=val
    return session_json



if __name__=="__main__":
    well_details=read_conns()
    print(well_details["312D"])
    for r in well_details["312D"]["routes"]:
        print(r["os"])
    # print(well_details["9834-1"]["comingled"])
    #
    # well_pcs=read_pcs()
    # print(well_pcs["15"])

    # well_maps=read_maps()
    # print(well_maps["9832"]["map"])
