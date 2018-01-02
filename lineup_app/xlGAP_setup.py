# -*- coding: utf-8 -*-
"""
Created on Mon Apr 25 14:34:45 2016

@author: zhumbo
"""

import numpy as np
# from lineup_app import PetexRoutines as PE
#from lineup_app import utils as ut
from openpyxl import load_workbook
import xlwings as xw
import os

#just testing github - deleteme23

def xlgap_conn():
    # get current directory using os library
    dirname, filename = os.path.split(os.path.abspath(__file__))
    # construct excel file full path
    xl_fullpath=os.path.join(dirname,r'static\xlgap.xlsx')
    # load excel file
    # xlgap_wb = xw.Workbook(xl_fullpath)
    xlgap_wb = xw.Book(xl_fullpath)
    # load sheet
    xlwells=xlgap_wb.sheets['Wells']
    xlunits=xlgap_wb.sheets['Units']
    # get max row number. Needed for furture looping through sheet rows
    #row_cnt = conns.max_row
    return xlwells,xlunits,xlgap_wb





def xl_get_well_data(unit,unit_id,well_data):

    xlwells,xlunits,xlgap_wb=xlgap_conn()
    wellname=[]
    gor=[]
    xlroutes=[]
    r=2
    while xlwells.range((r,1)).value!=None:
        if xlwells.range((r,3)).value==unit:
            wellname.append(wellname_type(xlwells.range((r,1)).value)) # to avoid decimal point coming from excel
            gor.append(xlwells.range((r,7)).value)
            xlroutes.append(str(xlwells.range((r,2)).value))
        r+=1

    for d,w in enumerate(wellname):

        this_route=-1
        route_cnt=0
        for ri,r in enumerate(well_data[w]["connection"]["routes"]):
            if r["os"] == xlroutes[d]:
                this_route=ri
                route_cnt+=1

        well_data[w]["gor"]=round(gor[d],1)
        well_data[w]["unit_id"]=unit_id
        well_data[w]["curr_route"]=this_route
        well_data[w]["route_cnt"]=route_cnt


    return well_data



def xl_get_all_well_data(well_data):

    """ SEQUENCE OF UNITS ====================================== """
    units=["KPC","U3","U2"]
    # units_simple=["kpc","u3","u2"]

    for idx,unit in enumerate(units):
        well_data=xl_get_well_data(unit,idx,well_data)

    return well_data



def xl_set_unit_routes(well_data):

    xlwells,xlunits,xlgap_wb=xlgap_conn()

    wellnames={}
    r=2
    while xlwells.range((r,1)).value is not None: # load wellnames in dict due to slow loops with excel
        wn=wellname_type(xlwells.range((r,1)).value)
        wellnames[wn]=r
        r+=1

    for well,val in well_data.items():
        # map selected_route to OS string
        route_open=""
        for i,r in enumerate(well_data[well]["connection"]["routes"]):
            route=str(r["unit"])+"--"+str(r["rms"])+"--"+str(r["tl"])+"--slot "+str(r["slot"])
            if route==val["selected_route"]:
                route_open=r["os"]

        wn_row=wellnames[well]
        xlwells.range((wn_row,2)).value=route_open

    xlgap_wb.save()

    return None


def xl_set_sep_pres(sep):

    xlwells,xlunits,xlgap_wb=xlgap_conn()

    xlunits.range((2,4)).value=sep["kpc_sep_pres"]
    xlunits.range((3,4)).value=sep["u2_sep_pres"]
    xlunits.range((4,4)).value=sep["u3_train1_sep_pres"]

    xlgap_wb.save()

    return None


def xl_get_sep_pres():

    xlwells,xlunits,xlgap_wb=xlgap_conn()

    sep={}
    sep["kpc_sep_pres"]=xlunits.range((2,4)).value
    sep["u2_sep_pres"]=xlunits.range((3,4)).value
    sep["u3_train1_sep_pres"]=xlunits.range((4,4)).value
    sep["u3_train2_sep_pres"]=xlunits.range((4,4)).value
    sep["u3_train3_sep_pres"]=xlunits.range((4,4)).value
    sep["u3_train4_sep_pres"]=xlunits.range((4,4)).value

    return sep


def wellname_type(wn):
    if type(wn) is not str: # to avoid decimal point coming from excel wellnames
        wn=int(wn)
    return str(wn)


if __name__=="__main__":
    import os
    data=xl_get_well_data("U2",1)
    alldata=xl_get_all_well_data()

    session_4xl={
        "15":{
            "fwhp":130,
            "in_opt":1,
            "selected_route":"U2--DIRECT--TL1--slot GORline"
        }
    }


    well_details={
        "15":{
            "routes":[{
                "rms":"DIRECT",
                "unit":"U2",
                "tl":"TL1",
                "os":"15_U2_L1",
                "slot":"GORline"
            }]
        }
    }

    xl_set_unit_routes(well_details,session_4xl)
    # print(data)
